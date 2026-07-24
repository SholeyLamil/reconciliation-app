"""
Streamlit app — Pelpay ↔ Gateway Settlement Reconciliation

Upload all files — auto-detects Pelpay, Cybersource, ChoicePay, and MPGS.

Usage:
    python -m streamlit run app.py
"""
import sys, os, tempfile, csv, io
from datetime import date, datetime
from collections import defaultdict
import streamlit as st
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from reconcile_core import run, DEFAULT_SCHEMAS, detect_date_range, _find_header_row

st.set_page_config(page_title='Reconciliation App', layout='wide')
st.title('Gateway ↔ Pelpay Settlement Reconciliation')

# ── Detection ────────────────────────────────────────────
def classify_file(headers):
    """Return ('PELPAY', None, None), ('SETTLEMENT', gateway, currency), or None."""
    h = [str(c).strip().lower() for c in headers]
    has = lambda s: any(s in col for col in h)

    # Pelpay: has Processor Reference + Merchant Name
    if has('processor reference') and has('merchant name'):
        # Check settlement gateway signatures too — some files might overlap
        if not has('merchant_ref_number') and not has('order reference'):
            return ('PELPAY', None, None)

    # Cybersource
    if has('merchant_ref_number') and has('amount') and has('merchant_id'):
        # Determine currency
        cur = None
        for col_h in ['currency']:
            if any(col_h == hh for hh in h):
                ci = h.index(col_h)
                # We'll read currency from data later
                cur = col_h
        return ('SETTLEMENT', 'CYBERSOURCE', cur)

    # ChoicePay / MPGS
    if has('order reference') and has('order amount') and has('merchant id'):
        cur = None
        for col_h in ['order amount (currency only)', 'currency']:
            if any(col_h == hh for hh in h):
                cur = col_h
        return ('SETTLEMENT', 'CHOICEPAY', cur)

    # MPGS
    if has('processor reference') and has('settlement amount') and has('merchant code'):
        cur = None
        for col_h in ['currency']:
            if any(col_h == hh for hh in h):
                cur = col_h
        return ('SETTLEMENT', 'MPGS', cur)

    return None

def read_currencies_from_file(fpath, cur_col_header, header_row=1):
    """Scan all data rows and return set of currencies found."""
    currencies = set()
    ext = os.path.splitext(fpath)[1].lower()
    try:
        if ext == '.csv':
            with open(fpath, encoding='utf-8-sig', newline='') as fh:
                reader = csv.reader(fh)
                h = [str(c).strip().lower() for c in next(reader, [])]
                if cur_col_header in h:
                    ci = h.index(cur_col_header)
                    for row in reader:
                        if row and len(row) > ci and row[ci]:
                            currencies.add(str(row[ci]).strip().upper())
        else:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            ws = wb[wb.sheetnames[0]]
            h = [str(c.value).strip().lower() for c in ws[header_row]]
            if cur_col_header in h:
                ci = h.index(cur_col_header)
                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    if row and len(row) > ci and row[ci]:
                        currencies.add(str(row[ci]).strip().upper())
            wb.close()
    except Exception:
        pass
    return currencies

def currency_from_filename(fpath):
    name = os.path.basename(fpath).upper()
    if 'NGN' in name and 'USD' not in name:
        return 'NGN'
    if 'USD' in name:
        return 'USD'
    return None

# ── Sidebar — status only ───────────────────────────────
st.sidebar.header('Status')
detected_range = st.sidebar.empty()  # filled after auto-detection

# ── Main — upload everything ─────────────────────────────
st.subheader('Upload Files')
st.markdown('Upload all daily files — the app auto-detects Pelpay vs settlement files and their currencies.')

uploaded_files = st.file_uploader(
    'Choose .xlsx or .csv files',
    type=['xlsx', 'csv'], accept_multiple_files=True,
    label_visibility='collapsed',
)

# ── Process uploads ─────────────────────────────────────-
if uploaded_files:
    pelpay_file = None
    settlements = []  # list of dicts

    for f in uploaded_files:
        ext = os.path.splitext(f.name)[1].lower()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext if ext else '.xlsx')
        tmp.write(f.getbuffer())
        tpath = tmp.name
        tmp.close()

        try:
            header_row = 1
            if ext == '.csv':
                with open(tpath, encoding='utf-8-sig', newline='') as fh:
                    reader = csv.reader(fh)
                    hdrs = next(reader, [])
            else:
                wb = openpyxl.load_workbook(tpath, data_only=True)
                ws = wb[wb.sheetnames[0]]
                hdrs, header_row = _find_header_row(ws)
                wb.close()

            cls = classify_file(hdrs)
            if cls is None:
                settlements.append({
                    'file': f, 'path': tpath, 'name': f.name,
                    'gateway': None, 'currency': None, 'status': 'UNKNOWN',
                })
                continue

            ftype, gw, cur_col = cls

            if ftype == 'PELPAY':
                if pelpay_file is None:
                    pelpay_file = {'file': f, 'path': tpath, 'name': f.name}
                else:
                    # Duplicate — mark as unknown
                    settlements.append({
                        'file': f, 'path': tpath, 'name': f.name,
                        'gateway': None, 'currency': None,
                        'status': 'DUPLICATE PELPAY (already have one)',
                    })
                continue

            # Settlement file — detect all currencies present
            currencies = set()
            if cur_col:
                currencies = read_currencies_from_file(tpath, cur_col, header_row)
            if not currencies:
                c = currency_from_filename(tpath)
                if c: currencies.add(c)
            if not currencies:
                currencies.add('?')

            for currency in sorted(currencies):
                settlements.append({
                    'file': f, 'path': tpath, 'name': f.name,
                    'gateway': gw, 'currency': currency, 'status': 'OK',
                })
        except Exception as e:
            settlements.append({
                'file': f, 'path': tpath, 'name': f.name,
                'gateway': None, 'currency': None,
                'status': f'Error: {e}',
            })

    # ── File checklist ────────────────────────────────────
    gateways = ['CYBERSOURCE', 'CHOICEPAY']
    gw_currencies = defaultdict(set)
    for s in settlements:
        if s['gateway'] and s['status'] == 'OK' and s['currency'] and s['currency'] != '?':
            gw_currencies[s['gateway']].add(s['currency'])

    st.subheader('Files Required')
    req_cols = st.columns(2)
    with req_cols[0]:
        st.markdown(f"{'✅' if pelpay_file else '❌'} **Pelpay file**")
    with req_cols[1]:
        st.markdown(f"{pelpay_file['name'] if pelpay_file else '— missing —'}")

    for gw in gateways:
        curs = gw_currencies.get(gw, set())
        ngn_ok = 'NGN' in curs
        usd_ok = 'USD' in curs
        label = {'CYBERSOURCE': 'Cybersource', 'CHOICEPAY': 'ChoicePay/MPGS'}[gw]
        status = '✅' if ngn_ok and usd_ok else '⚠️' if ngn_ok or usd_ok else '❌'
        cur_label = f"NGN {'✅' if ngn_ok else '❌'}, USD {'✅' if usd_ok else '❌'}"
        with req_cols[0]:
            st.markdown(f"{status} **{label}** — {cur_label}")
        with req_cols[1]:
            files = [s['name'] for s in settlements if s['gateway'] == gw and s['status'] == 'OK']
            st.caption(', '.join(set(files)) if files else '— missing —')

    # ── Detection table ───────────────────────────────────
    st.subheader('Detection Results')
    file_map = defaultdict(lambda: {'type': '—', 'gateway': '—', 'currencies': set(), 'status': 'OK'})
    if pelpay_file:
        file_map[pelpay_file['name']] = {'type': 'Pelpay', 'gateway': '—', 'currencies': set(), 'status': 'OK'}
    for s in settlements:
        fm = file_map[s['name']]
        fm['type'] = 'Settlement' if s['gateway'] else '—'
        fm['gateway'] = s['gateway'] if s['gateway'] else '—'
        if s['currency']: fm['currencies'].add(s['currency'])
        if s['status'] != 'OK': fm['status'] = s['status']
    rows = [{'File': k, 'Type': v['type'], 'Gateway': v['gateway'],
             'Currency': ', '.join(sorted(v['currencies'])) if v['currencies'] else '—',
             'Status': v['status']} for k, v in file_map.items()]
    st.table(rows)

    # ── Validation ───────────────────────────────────────
    # errors BLOCK the run; warnings are informational. A currency being absent
    # (e.g. the NGN merchant had issues that day) is a warning, not a blocker —
    # the app still reconciles whatever currencies/gateways are present.
    errors, warnings = [], []
    if not pelpay_file:
        errors.append('No Pelpay file detected. Must contain: Processor Reference, Merchant Name')
    valid_settlements = [s for s in settlements if s['gateway'] and s['status'] == 'OK' and s['currency'] and s['currency'] != '?']
    unknown = [s for s in settlements if not s['gateway'] and s['status'] not in ('DUPLICATE PELPAY (already have one)',)]
    if unknown:
        errors.append(f'{len(unknown)} file(s) could not be matched. '
                       'Cybersource files need: merchant_ref_number, amount, merchant_id. '
                       'ChoicePay files need: Order Reference, Order Amount, Merchant ID.')
    if not valid_settlements:
        errors.append('No settlement files with recognised data were detected.')
    for gw in gateways:
        curs = gw_currencies.get(gw, set())
        if 'NGN' not in curs and 'USD' not in curs:
            warnings.append(f'No {gw} file this run — reconciling without it.')
        elif 'NGN' not in curs:
            warnings.append(f'No NGN rows for {gw} this run — reconciling USD only for it.')
        elif 'USD' not in curs:
            warnings.append(f'No USD rows for {gw} this run — reconciling NGN only for it.')
    for s in settlements:
        if s['gateway'] and s['currency'] == '?':
            errors.append(f'Could not determine currency for {s["name"]}. Rename file to include NGN or USD.')

    for w in warnings:
        st.warning(w)
    for e in errors:
        st.error(e)

    st.divider()

    # ── Run ──────────────────────────────────────────────
    if not errors and pelpay_file:
        settle_items = [(s['gateway'], s['currency'], s['path']) for s in valid_settlements]

        # Auto-detect date range from settlement files
        try:
            dr = detect_date_range(settle_items)
            date_min, date_max = dr
            date_label = f"{date_min.strftime('%d %b %Y')} – {date_max.strftime('%d %b %Y')}"
            detected_range.info(f'Detected settlement range: {date_label}')
        except Exception as e:
            detected_range.error(f'Could not detect dates: {e}')
            dr = None

        if st.button('Run Reconciliation', type='primary'):
            if dr is None:
                st.error('Cannot run without a valid date range. Check settlement files.')
                st.stop()
            tmpdir = tempfile.mkdtemp()
            try:
                pel_path = pelpay_file['path']

                if date_min == date_max:
                    fname = f'Reconciliation_{date_min.strftime("%Y-%m-%d")}.xlsx'
                else:
                    fname = f'Reconciliation_{date_min.strftime("%Y-%m-%d")}_to_{date_max.strftime("%Y-%m-%d")}.xlsx'
                out_path = os.path.join(tmpdir, fname)

                with st.spinner('Running reconciliation…'):
                    result = run(pel_path, settle_items, dr, out_path)

                st.success('Reconciliation complete!')
                c1, c2, c3 = st.columns(3)
                c1.metric('Settlement Rows', result['settle_rows'])
                c2.metric('Matched', result['matched'])
                c3.metric('Unsettled Transactions', result['exceptions'])
                st.write('Sheets:', ', '.join(result['sheets']))

                with open(out_path, 'rb') as fh:
                    st.download_button(
                        label='⬇ Download Reconciliation Workbook',
                        data=fh,
                        file_name=os.path.basename(out_path),
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    )
            except Exception as e:
                st.error(f'Reconciliation failed: {e}')
                import traceback; st.code(traceback.format_exc())

else:
    st.info('Upload .xlsx or .csv files to get started.')

# ── Guide ────────────────────────────────────────────────
with st.expander('How to use'):
    st.markdown('''
 1. **Upload all .xlsx / .csv files** — Pelpay + settlement files, all at once.  
   The app auto-detects:
    - **Pelpay file** (needs: Processor Reference, Merchant Name)
    - **Cybersource files** (needs: merchant_ref_number, amount, merchant_id)
    - **ChoicePay files** (needs: Order Reference, Order Amount, Merchant ID)
    - **MPGS files** (needs: Processor Reference, Settlement Amount, Merchant Code)
    - **Currency** from file content or filename (NGN / USD)
    - **Settlement date range** from the files themselves
 2. Review the detection table and the auto-detected date range in the sidebar.
 3. Click **Run Reconciliation**.
 4. Download the output workbook.
    ''')
