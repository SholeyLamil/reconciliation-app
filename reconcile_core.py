"""
Core reconciliation engine — parameterized, importable, no I/O globals.
"""
import sys, os, tempfile
from collections import defaultdict, Counter
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Schemas ──────────────────────────────────────────────
DEFAULT_SCHEMAS = {
    'CYBERSOURCE': {
        'ref': 'merchant_ref_number', 'amount': 'amount',
        'merchant_field': 'merchant_id', 'currency': 'currency',
        'status': 'status', 'approved_statuses': {'BATCHED'},
        'date': 'batch_date',
    },
    'CHOICEPAY': {
        'ref': 'Order Reference', 'amount': 'Order Amount (amount only)',
        'merchant_field': 'Merchant ID', 'currency': 'Order Amount (currency only)',
        'status': 'Order Status', 'approved_statuses': {'Captured'},
        'date': 'Order Date',
    },
    'MPGS': {
        'ref': 'Processor Reference', 'amount': 'Settlement amount',
        'merchant_field': 'Merchant code', 'currency': 'Currency',
        'status': 'Transaction Status', 'approved_statuses': {'Successful'},
        'date': 'Transaction Date',
    },
}

PELPAY_ALIASES = {
    'ref': ['Processor Reference'], 'date': ['Transaction Date'],
    'amount': ['Amount Collected'], 'status': ['Transaction Status'],
    'merchant_name': ['Merchant Name'], 'merchant_code': ['Merchant code'],
    'currency': ['Currency'], 'merchant_reference': ['Merchant Reference'],
    'payment_reference': ['Payment Reference'],
}

PELPAY_FIELDS = ['Transaction Date', 'Payment Reference', 'Advice Id', 'Merchant Name',
    'Merchant code', 'Merchant Reference', 'Processor Reference', 'Currency',
    'Channel', 'Transaction Status', 'Gross Amount', 'Amount Collected',
    'Processing Fee Applied', 'Merchant Settlement']

# ── Styling constants ────────────────────────────────────
TITLE_FILL = PatternFill('solid', fgColor='17365D')
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
MATCH_FILL = PatternFill('solid', fgColor='E2F0D9')
MISSING_FILL = PatternFill('solid', fgColor='FCE4D6')
EXC_HEADER_FILL = PatternFill('solid', fgColor='F4B183')
TOTAL_FILL = PatternFill('solid', fgColor='D9EAF7')
WHITE_BOLD = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
BLACK_BOLD = Font(name='Calibri', size=11, bold=True)
AMT_FMT = '#,##0.00'

# ── Helpers ──────────────────────────────────────────────
def norm(x):
    return '' if x is None else str(x).strip().upper()

def ordinal(n):
    if 11 <= n % 100 <= 13: suffix = 'th'
    else: suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return f"{n}{suffix}"

def parse_dt(x):
    if x is None: return None
    if isinstance(x, datetime): return x
    if isinstance(x, date): return datetime.combine(x, datetime.min.time())
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%SZ', '%Y-%m-%d',
                '%m/%d/%Y %H:%M', '%m/%d/%Y', '%d/%m/%Y %H:%M', '%d/%m/%Y'):
        try: return datetime.strptime(str(x).strip(), fmt)
        except Exception: pass
    return None

def resolve_headers(headers, aliases, label):
    idx = {h: i for i, h in enumerate(headers)}
    resolved = {}
    for field, options in aliases.items():
        found = next((o for o in options if o in idx), None)
        if found is None:
            raise ValueError(f"Missing '{field}' in {label}. Available: {list(idx)}")
        resolved[field] = found
    return resolved, idx

def autosize(ws, ncols, minwidth=14, maxwidth=40):
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        maxlen = max((len(str(cell.value or '')) for row in ws.iter_rows(min_col=c, max_col=c) for cell in row), default=0)
        ws.column_dimensions[letter].width = max(minwidth, min(maxwidth, maxlen + 2))

def write_title_row(ws, r, ncols, value):
    ws.cell(r, 1, value=value).font = WHITE_BOLD
    for c in range(1, ncols + 1):
        ws.cell(r, c).fill = TITLE_FILL

def write_header_row(ws, r, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.fill, cell.font = HEADER_FILL, WHITE_BOLD

def write_total_row(ws, r, ncols, label=None):
    if label: ws.cell(r, 1, label)
    for c in range(1, ncols + 1):
        ws.cell(r, c).fill = TOTAL_FILL
        ws.cell(r, c).font = BLACK_BOLD

def apply_formatting(wb):
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                elif isinstance(cell.value, float):
                    cell.number_format = AMT_FMT
                elif isinstance(cell.value, int) and abs(cell.value) >= 100:
                    cell.number_format = AMT_FMT

# ── Auto-detect date range from settlement files ────────
def detect_date_range(settlement_files, schemas=None):
    """Scan settlement files for their date columns and return (min_date, max_date).
    settlement_files: list of (gateway_name, currency, file_path)
    Returns (date, date) tuple or raises ValueError if no dates can be parsed."""
    if schemas is None:
        schemas = DEFAULT_SCHEMAS
    all_dates = []
    for gw, cur, fpath in settlement_files:
        schema = schemas[gw]
        s_headers, s_rows = load_file_rows(fpath)
        s_idx = {h: i for i, h in enumerate(s_headers)}
        date_col = schema.get('date')
        if date_col and date_col in s_idx:
            di = s_idx[date_col]
        else:
            for alt in ('batch_date', 'date', 'Order Date', 'transaction_date', 'Transaction Date'):
                if alt in s_idx:
                    di = s_idx[alt]
                    break
            else:
                continue
        for row in s_rows:
            d = _parse_cell_date(row[di])
            if d: all_dates.append(d)
    if not all_dates:
        raise ValueError('Could not detect any dates from settlement files.')
    return (min(all_dates), max(all_dates))

# ── Core reconciliation ──────────────────────────────────
def load_pelpay(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers, hr = _find_header_row(ws)
    res, idx = resolve_headers(headers, PELPAY_ALIASES, 'Pelpay')
    rows = [r for r in ws.iter_rows(min_row=hr + 1, values_only=True) if tuple(r) != tuple(headers)]
    wb.close()

    pel_by_ref = {}
    for r in rows:
        ref = r[idx[res['ref']]]
        if ref is None: continue
        dt = parse_dt(r[idx[res['date']]])
        pel_by_ref[norm(ref)] = {
            'ref': norm(ref), 'dt': dt,
            'amount': r[idx[res['amount']]] or 0,
            'status': str(r[idx[res['status']]] or '').strip(),
            'merchant': str(r[idx[res['merchant_name']]] or '').strip(),
            'currency': str(r[idx[res['currency']]] or '').strip(),
            'merchant_ref': r[idx[res['merchant_reference']]],
            'payment_ref': r[idx[res['payment_reference']]],
            'gross': r[idx['Gross Amount']] if 'Gross Amount' in idx else 0,
            'collected': r[idx['Amount Collected']] if 'Amount Collected' in idx else 0,
            'merchant_settlement': r[idx['Merchant Settlement']] if 'Merchant Settlement' in idx else 0,
            'merchant_code': r[idx[res['merchant_code']]],
            'advice_id': r[idx['Advice Id']] if 'Advice Id' in idx else None,
            'channel': r[idx['Channel']] if 'Channel' in idx else None,
            'payment_ref_raw': r[idx[res['payment_reference']]],
            'fee': r[idx['Processing Fee Applied']] if 'Processing Fee Applied' in idx else None,
        }
    return pel_by_ref

def in_date_range(dt, date_range):
    """Check if a datetime falls within a date range. date_range is (start, end) or a single date."""
    if dt is None: return False
    d = dt.date() if isinstance(dt, datetime) else dt
    if isinstance(date_range, tuple):
        return date_range[0] <= d <= date_range[1]
    return d == date_range

def _find_header_row(ws):
    """Scan worksheet rows to find the actual header row, skipping metadata rows."""
    for rnum in range(1, min(ws.max_row + 1, 11)):
        row = [ws.cell(rnum, c).value for c in range(1, ws.max_column + 1)]
        non_empty = sum(1 for v in row if v is not None and str(v).strip())
        if non_empty < 3:
            continue
        combined = ' '.join(str(v or '') for v in row)
        if any(kw in combined for kw in ['chamswitch', 'Report_', 'Daily_Classic']):
            continue
        return [str(v).strip() if v is not None else '' for v in row], rnum
    raise ValueError('Could not find header row in file')

def _find_header_row_list(all_rows):
    """Find the real header row in a list of CSV rows, skipping metadata rows."""
    for i, row in enumerate(all_rows[:10]):
        non_empty = sum(1 for v in row if v is not None and str(v).strip())
        if non_empty < 3:
            continue
        combined = ' '.join(str(v or '') for v in row)
        if any(kw in combined for kw in ['chamswitch', 'Report_', 'Daily_Classic']):
            continue
        return [str(v).strip() if v is not None else '' for v in row], i
    raise ValueError('Could not find header row in file')

def _num_or_str(v):
    """Parse a cell to float, tolerating thousands separators (e.g. '16,072.25'); else str."""
    s = v.strip()
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return s

def load_file_rows(fpath):
    """Load headers and data rows from .xlsx or .csv. Returns (headers, rows)."""
    if fpath.lower().endswith('.csv'):
        import csv
        with open(fpath, 'r', encoding='utf-8-sig', newline='') as f:
            all_rows = list(csv.reader(f))
        headers, hi = _find_header_row_list(all_rows)
        rows = []
        for row in all_rows[hi + 1:]:
            if not any(str(c).strip() for c in row):
                continue
            parsed = [(_num_or_str(v) if (v is not None and v.strip()) else v) for v in row]
            rows.append(tuple(parsed))
        return headers, rows
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers, hr = _find_header_row(ws)
    rows = [r for r in ws.iter_rows(min_row=hr + 1, values_only=True) if tuple(r) != tuple(headers)]
    wb.close()
    return headers, rows

def load_settlements(pel_by_ref, settlement_files, date_range, schemas=None):
    """settlement_files: list of (gateway_name, currency, file_path)
       date_range: date or (start, end) tuple — used for exception date filtering"""
    if schemas is None:
        schemas = DEFAULT_SCHEMAS

    # Phase 1: load + group by merchant_id + resolve display name
    raw_sections = []
    for gw, cur, fpath in settlement_files:
        schema = schemas[gw]
        s_headers, s_rows = load_file_rows(fpath)
        s_idx = {h: i for i, h in enumerate(s_headers)}

        approved = [r for r in s_rows if r[s_idx[schema['status']]] in schema['approved_statuses']]
        grp = defaultdict(list)
        use_file_currency = schema.get('currency') and schema['currency'] in s_headers
        for r in approved:
            mid = r[s_idx[schema['merchant_field']]]
            if use_file_currency:
                raw_cur = r[s_idx[schema['currency']]]
                file_cur = str(raw_cur).strip().upper() if raw_cur else 'NONE'
                grp[(mid, file_cur)].append(r)
            else:
                grp[mid].append(r)

        for key, grp_rows in grp.items():
            if use_file_currency:
                raw_key, file_cur = key
                cur = file_cur
            else:
                raw_key, file_cur = key, cur
            names = Counter()
            for r in grp_rows:
                p = pel_by_ref.get(norm(r[s_idx[schema['ref']]]))
                if p and p['merchant']:
                    names[p['merchant'].title()] += 1
            display = names.most_common(1)[0][0] if names else str(raw_key).title()

            raw_sections.append({
                'gw': gw, 'currency': cur, 'merchant': display,
                'merchant_code': raw_key, 'headers': s_headers, 'schema': schema,
                'rows': grp_rows, 'settle_refs': {norm(r[s_idx[schema['ref']]]) for r in grp_rows},
                's_idx': s_idx,
            })

    # Phase 2: merge by (merchant, currency, gw) — combine rows, dedup by ref
    merged = {}
    for s in raw_sections:
        key = (s['merchant'], s['currency'], s['gw'])
        if key in merged:
            m = merged[key]
            seen = {norm(r[m['headers'].index(m['schema']['ref'])]) for r in m['rows']}
            for r in s['rows']:
                ref = norm(r[s['headers'].index(s['schema']['ref'])])
                if ref not in seen:
                    m['rows'].append(r); seen.add(ref)
            m['settle_refs'].update(s['settle_refs'])
        else:
            merged[key] = dict(s)

    # Phase 3: match + compute exceptions per merged group.
    # Pool settled refs across ALL gateways for each (merchant, currency): a txn
    # settled under one gateway must not be flagged "missing" by another gateway.
    settled_by_mc = defaultdict(set)
    for s in merged.values():
        settled_by_mc[(s['merchant'], s['currency'])].update(s['settle_refs'])

    exc_seen = set()  # a genuinely-missing ref is reported once, not per gateway
    sections = []
    for key, s in merged.items():
        schema = s['schema']
        s_idx = s['s_idx']
        s_headers = s['headers']
        matched, unmatched, exc = [], [], []
        all_refs = settled_by_mc[(s['merchant'], s['currency'])]

        for r in s['rows']:
            ref = norm(r[s_idx[schema['ref']]])
            p = pel_by_ref.get(ref)
            if p and p['merchant'].title() == s['merchant'] and p['currency'] == s['currency']:
                diff = r[s_idx[schema['amount']]] - p['amount']
                matched.append((r, p, diff, s_idx, schema))
            else:
                unmatched.append(r)

        for p in pel_by_ref.values():
            if (p['status'] == 'Successful' and p['currency'] == s['currency']
                and p['merchant'].title() == s['merchant']
                and p['dt'] and in_date_range(p['dt'], date_range)
                and p['ref'] not in all_refs
                and p['ref'] not in exc_seen):
                exc.append(p)
                exc_seen.add(p['ref'])

        sections.append({
            'gw': s['gw'], 'currency': s['currency'], 'merchant': s['merchant'],
            'merchant_code': s['merchant_code'], 'headers': s_headers, 'schema': schema,
            's_idx': s_idx,
            'rows': s['rows'], 'matched': matched, 'unmatched': unmatched,
            'exceptions': exc,
        })
    return sections

def date_label(d):
    if isinstance(d, tuple):
        start, end = d
        if start == end:
            return f"{ordinal(end.day)} {end.strftime('%B')} {end.strftime('%Y')}"
        return f"{ordinal(start.day)} {start.strftime('%B')} – {ordinal(end.day)} {end.strftime('%B')} {end.strftime('%Y')}"
    return f"{ordinal(d.day)} {d.strftime('%B')} {d.strftime('%Y')}"

# ── Workbook builders ────────────────────────────────────
def _resolve_date_col(s_headers, schema):
    """Find the date column in headers: prefer schema date, fallback to common names."""
    preferred = schema.get('date')
    if preferred and preferred in s_headers:
        return s_headers.index(preferred)
    for alt in ('batch_date', 'date', 'Order Date', 'transaction_date', 'Transaction Date'):
        if alt in s_headers:
            return s_headers.index(alt)
    return None

def _parse_cell_date(val):
    """Parse a cell value to a date object."""
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    if val:
        raw = str(val).strip()
        for fmt in ('%Y-%m-%dT%H:%M:%SZ', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d',
                    '%B %d, %Y', '%m/%d/%Y %H:%M', '%m/%d/%Y', '%d/%m/%Y %H:%M', '%d/%m/%Y'):
            try: return datetime.strptime(raw, fmt).date()
            except: pass
    return None

def get_row_dates(sec):
    """Return dict mapping settlement row ref → date for a section."""
    schema = sec['schema']
    s_headers = sec['headers']
    di = _resolve_date_col(s_headers, schema)
    if di is None:
        return {}
    result = {}
    for r in sec['rows']:
        ref = norm(r[s_headers.index(schema['ref'])])
        result[ref] = _parse_cell_date(r[di])
    return result

def _row_date_val(row, s_idx, schema):
    """Extract and parse date from a settlement row using the schema's date field."""
    preferred = schema.get('date')
    di = None
    if preferred and preferred in s_idx:
        di = s_idx[preferred]
    else:
        for alt in ('batch_date', 'date', 'Order Date', 'transaction_date', 'Transaction Date'):
            if alt in s_idx:
                di = s_idx[alt]
                break
    if di is not None:
        return _parse_cell_date(row[di])
    return None

def daily_stats_for_sec(sec, d, pel_by_ref):
    """Compute summary stats for one sec on one date d."""
    row_dates = get_row_dates(sec)
    s_headers, schema = sec['headers'], sec['schema']
    s_idx = sec['s_idx']

    matched_today = [m for m in sec['matched']
                     if _row_date_val(m[0], m[3], m[4]) == d]
    exc_today = [e for e in sec['exceptions']
                 if e['dt'] and e['dt'].date() == d]
    settle_rows_today = sum(1 for r in sec['rows']
                            if row_dates.get(norm(r[s_headers.index(schema['ref'])])) == d)

    settle_total = sum(r[s_idx[schema['amount']]] for r in sec['rows']
                       if row_dates.get(norm(r[s_headers.index(schema['ref'])])) == d)
    pelpay_total = sum(m[1]['amount'] for m in matched_today)
    diff_total = settle_total - pelpay_total
    exc_amt = sum(e['amount'] for e in exc_today)

    return {
        'settle_rows': settle_rows_today,
        'matched_rows': len(matched_today),
        'settle_total': settle_total,
        'pelpay_total': pelpay_total,
        'diff_total': diff_total,
        'exc_count': len(exc_today),
        'exc_amt': exc_amt,
    }

def build_summary(wb, sections, date_range, pel_by_ref):
    ws = wb.create_sheet('SUMMARY', 0)
    label = date_label(date_range)
    is_range = isinstance(date_range, tuple)
    end_date = date_range[1] if is_range else date_range
    days = [date_range[0] + __import__('datetime').timedelta(days=i)
            for i in range((end_date - date_range[0]).days + 1)] if is_range else [date_range]

    ws.cell(1, 1, value=f'Gateway \u2194 Pelpay Settlement Reconciliation \u2013 Summary').font = Font(bold=True, size=13)
    ws.cell(2, 1, value=f'Settlement Date: {label}').font = Font(italic=True)

    sum_headers = [
        'Settlement Date', 'Currency', 'Settlement Rows', 'Matched Rows',
        'Settlement Total', 'Matched Pelpay Total', 'Difference',
        'Missing (Pelpay)', 'Missing Amount (Pelpay)', 'Coverage', 'Gateway',
    ]

    grand = {cur: defaultdict(float) for cur in ['NGN', 'USD']}
    r = 5

    for cur in ['NGN', 'USD']:
        merch_sections = sorted(
            [s for s in sections if s['currency'] == cur],
            key=lambda x: x['merchant']
        )
        for sec in merch_sections:
            sec_grand = defaultdict(float)
            dated_refs = set()
            row_dates = get_row_dates(sec)
            s_headers, schema, s_idx = sec['headers'], sec['schema'], sec['s_idx']

            for d in days:
                st = daily_stats_for_sec(sec, d, pel_by_ref)
                for gk in ['settle_rows', 'matched_rows', 'settle_total', 'pelpay_total',
                            'diff_total', 'exc_count', 'exc_amt']:
                    sec_grand[gk] += st[gk]
                    grand[cur][gk] += st[gk]
                for sr in sec['rows']:
                    ref = norm(sr[s_headers.index(schema['ref'])])
                    if row_dates.get(ref) == d:
                        dated_refs.add(ref)

            # Catch rows whose dates fall outside the settlement date range
            other_rows = [sr for sr in sec['rows'] if norm(sr[s_headers.index(schema['ref'])]) not in dated_refs]
            other_matched_refs = {norm(m[0][m[3][m[4]['ref']]]) for m in sec['matched']} - dated_refs
            other_matched_count = len(other_matched_refs)
            if other_rows:
                other_settle_amt = sum(sr[s_idx[schema['amount']]] for sr in other_rows)
                other_pelpay_amt = sum(m[1]['amount'] for m in sec['matched']
                                       if norm(m[0][m[3][m[4]['ref']]]) in other_matched_refs)
                for gk, val in [('settle_rows', len(other_rows)), ('matched_rows', other_matched_count),
                                 ('settle_total', other_settle_amt), ('pelpay_total', other_pelpay_amt),
                                 ('diff_total', other_settle_amt - other_pelpay_amt)]:
                    sec_grand[gk] += val
                    grand[cur][gk] += val

            if not sec_grand.get('settle_rows') and not sec_grand.get('matched_rows'):
                continue

            ws.cell(r, 1, value=sec['merchant']).font = Font(bold=True, size=12)
            r += 1
            write_header_row(ws, r, sum_headers); r += 1

            if other_rows:
                vals = ['Other Dates', cur, len(other_rows), other_matched_count,
                        other_settle_amt, other_pelpay_amt,
                        other_settle_amt - other_pelpay_amt,
                        '', '', 'PARTIAL', sec['gw']]
                for c, v in enumerate(vals, 1): ws.cell(r, c, v)
                r += 1

            for d in days:
                st = daily_stats_for_sec(sec, d, pel_by_ref)
                if not st['settle_rows'] and not st['matched_rows']:
                    continue

                vals = [d.isoformat(), cur, st['settle_rows'], st['matched_rows'],
                        st['settle_total'], st['pelpay_total'], st['diff_total'],
                        st['exc_count'], st['exc_amt'], 'PARTIAL', sec['gw']]
                for c, v in enumerate(vals, 1): ws.cell(r, c, v)
                r += 1

            if sec_grand:
                vals = [f'{cur} TOTAL', None, sec_grand['settle_rows'], sec_grand['matched_rows'],
                        sec_grand['settle_total'], sec_grand['pelpay_total'], sec_grand['diff_total'],
                        sec_grand['exc_count'], sec_grand['exc_amt'], None, None]
                write_total_row(ws, r, 9)
                for c, v in enumerate(vals, 1): ws.cell(r, c, v)
                r += 2

    for cur in ['NGN', 'USD']:
        g = grand[cur]
        if not g.get('settle_rows', 0) and not g.get('matched_rows', 0): continue
        vals = [f'{cur} GRAND TOTAL', None, g['settle_rows'], g['matched_rows'],
                g['settle_total'], g['pelpay_total'], g['diff_total'],
                g['exc_count'], g['exc_amt'], None, None]
        write_total_row(ws, r, 9)
        for c, v in enumerate(vals, 1): ws.cell(r, c, v)
        r += 1

    r += 1
    ws.cell(r, 1, value='Successful Pelpay Transaction Summary').font = BLACK_BOLD; r += 1
    ps_headers = ['Merchant', 'Currency', 'Successful Count', 'Gross Total',
                   'Amount Collected Total', 'Merchant Settlement Total']
    write_header_row(ws, r, ps_headers); r += 1

    for cur in ['NGN', 'USD']:
        for sec in sorted(sections, key=lambda x: x['merchant']):
            if sec['currency'] != cur: continue
            succ_rows = [p for p in pel_by_ref.values()
                         if p['status'] == 'Successful' and p['currency'] == cur
                         and p['merchant'].title() == sec['merchant']
                         and p['dt'] and in_date_range(p['dt'], date_range) and p['ref']]
            if not succ_rows: continue
            ws.cell(r, 1, sec['merchant']); ws.cell(r, 2, cur)
            ws.cell(r, 3, len(succ_rows))
            ws.cell(r, 4, sum(p['gross'] or 0 for p in succ_rows))
            ws.cell(r, 5, sum(p['collected'] or 0 for p in succ_rows))
            ws.cell(r, 6, sum(p['merchant_settlement'] or 0 for p in succ_rows))
            r += 1
    autosize(ws, len(sum_headers))

def build_received(wb, sections, date_range, pel_by_ref, all_merchant_cur):
    ws = wb.create_sheet('All received Transactions')
    label = date_label(date_range)
    ncols = len(PELPAY_FIELDS) + 1
    r = 1

    write_title_row(ws, r, ncols, f'All Received Successful Transactions \u2013 {label}')
    r += 1

    grand_rec = defaultdict(lambda: {'count': 0, 'amount': 0.0})
    for cur in ['NGN', 'USD']:
        for merchant, currency in all_merchant_cur:
            if currency != cur: continue
            rows = [p for p in pel_by_ref.values()
                    if p['status'] == 'Successful' and p['currency'] == cur
                    and p['merchant'].title() == merchant
                    and p['dt'] and in_date_range(p['dt'], date_range) and p['ref']]
            grand_rec[cur]['count'] += len(rows)
            grand_rec[cur]['amount'] += sum(p['amount'] for p in rows)
    write_total_row(ws, r, ncols, 'Grand Total')
    ws.cell(r, 13, grand_rec['NGN']['amount'] + grand_rec['USD']['amount'])
    r += 2

    for cur in ['NGN', 'USD']:
        for merchant, currency in all_merchant_cur:
            if currency != cur: continue
            day_rows = sorted(
                [p for p in pel_by_ref.values()
                 if p['status'] == 'Successful' and p['currency'] == cur
                 and p['merchant'].title() == merchant
                 and p['dt'] and in_date_range(p['dt'], date_range)],
                key=lambda x: x['dt'] or datetime.min)
            title = f'ALL RECEIVED SUCCESSFUL TRANSACTIONS \u2014 {merchant} \u2014 {currency} \u2014 {label}'
            write_title_row(ws, r, ncols, title); r += 1
            day_total = sum(p['amount'] for p in day_rows)
            write_total_row(ws, r, ncols, 'Total')
            ws.cell(r, 13, day_total); ws.cell(r, 14, len(day_rows)); r += 1
            ws.cell(r, 1, 'received_date')
            for c, h in enumerate(PELPAY_FIELDS, 2): ws.cell(r, c, h)
            for c in range(1, ncols + 1): ws.cell(r, c).fill = HEADER_FILL; ws.cell(r, c).font = WHITE_BOLD
            r += 1
            if day_rows:
                for p in day_rows:
                    ws.cell(r, 1, p['dt'].date().isoformat() if p['dt'] else ''); ws.cell(r, 2, p['dt'])
                    ws.cell(r, 3, p['payment_ref']); ws.cell(r, 4, p['advice_id'])
                    ws.cell(r, 5, p['merchant']); ws.cell(r, 6, p['merchant_code'])
                    ws.cell(r, 7, p['merchant_ref']); ws.cell(r, 8, p['payment_ref_raw'])
                    ws.cell(r, 9, p['currency']); ws.cell(r, 10, p['channel'])
                    ws.cell(r, 11, p['status']); ws.cell(r, 12, p['gross'])
                    ws.cell(r, 13, p['collected']); ws.cell(r, 14, p['fee'])
                    ws.cell(r, 15, p['merchant_settlement']); r += 1
            else:
                ws.cell(r, 1, 'No transactions found.'); r += 1
            write_total_row(ws, r, ncols, 'Total')
            ws.cell(r, 13, day_total); ws.cell(r, 14, len(day_rows))
            r += 2
    autosize(ws, ncols)

def build_missing_sheets(wb, sections, date_range, all_merchant_cur):
    """Create two sheets:
       'Missing from Settlement' — Pelpay transactions not found in settlement
       'Missing from Pelpay'    — Settlement rows not found in Pelpay
    """
    label = date_label(date_range)
    exc_fields = ['received_date', 'pelpay_processor_reference', 'pelpay_transaction_date',
                   'pelpay_transaction_amount', 'pelpay_status', 'pelpay_merchant_reference',
                   'pelpay_payment_reference']

    # ── Sheet 1: Missing from Settlement (Pelpay-only) ─────────────────
    ws = wb.create_sheet('Missing from Settlement')
    r = 1
    write_title_row(ws, r, 7, f'In Pelpay But Not Found in Settlement \u2013 {label}')
    r += 2

    for cur in ['NGN', 'USD']:
        for merchant, currency in all_merchant_cur:
            if currency != cur: continue
            day_exc = [e for e in sum((s['exceptions'] for s in sections
                                        if s['merchant'] == merchant and s['currency'] == currency), [])
                       if e['dt'] and in_date_range(e['dt'], date_range)]
            title = f'{merchant} \u2014 {currency} \u2014 {label}'
            ws.cell(r, 1, title)
            for c in range(1, 8): ws.cell(r, c).fill = EXC_HEADER_FILL; ws.cell(r, c).font = BLACK_BOLD
            r += 1
            exc_amt = sum(e['amount'] for e in day_exc)
            for c, val in [(1, 'Missing Count'), (2, len(day_exc)),
                           (3, 'Total Amount'), (4, exc_amt)]: ws.cell(r, c, val)
            for c in range(1, 5): ws.cell(r, c).fill = TOTAL_FILL; ws.cell(r, c).font = BLACK_BOLD
            r += 1
            for c, h in enumerate(exc_fields, 1):
                cell = ws.cell(r, c, h); cell.fill, cell.font = EXC_HEADER_FILL, WHITE_BOLD
            r += 1
            if day_exc:
                for e in sorted(day_exc, key=lambda x: x['dt'] or datetime.min):
                    ws.cell(r, 1, e['dt'].date().isoformat() if e['dt'] else ''); ws.cell(r, 2, e['payment_ref_raw'])
                    ws.cell(r, 3, e['dt']); ws.cell(r, 4, e['amount'])
                    ws.cell(r, 5, e['status']); ws.cell(r, 6, e['merchant_ref'])
                    ws.cell(r, 7, e['payment_ref'])
                    for c in range(1, 8): ws.cell(r, c).fill = MISSING_FILL
                    r += 1
            else:
                ws.cell(r, 1, 'All transactions matched.'); r += 1
            for c, val in [(1, 'Missing Count'), (2, len(day_exc)),
                           (3, 'Total Amount'), (4, exc_amt)]: ws.cell(r, c, val)
            for c in range(1, 5): ws.cell(r, c).fill = TOTAL_FILL; ws.cell(r, c).font = BLACK_BOLD
            r += 2
    autosize(ws, 7)

    # ── Sheet 2: Missing from Pelpay (Settlement-only) ────────────────
    ws2 = wb.create_sheet('Missing from Pelpay')
    r = 1
    write_title_row(ws2, r, 7, f'In Settlement But Not Found in Pelpay \u2013 {label}')
    r += 2

    un_fields = ['settlement_ref', 'settlement_date', 'settlement_amount', 'settlement_currency',
                 'merchant_id', 'gateway', 'merchant_name']

    for cur in ['NGN', 'USD']:
        for merchant, currency in all_merchant_cur:
            if currency != cur: continue
            sec = next((s for s in sections if s['merchant'] == merchant and s['currency'] == currency), None)
            if not sec: continue
            unmatched = sec.get('unmatched', [])
            if not unmatched: continue
            s_headers, schema = sec['headers'], sec['schema']
            s_i = {h: i for i, h in enumerate(s_headers)}

            title = f'{merchant} \u2014 {currency} \u2014 {label}'
            ws2.cell(r, 1, title)
            for c in range(1, 8): ws2.cell(r, c).fill = EXC_HEADER_FILL; ws2.cell(r, c).font = BLACK_BOLD
            r += 1

            un_amt = sum(u[s_i[schema['amount']]] for u in unmatched)
            for c, val in [(1, 'Missing Count'), (2, len(unmatched)),
                           (3, 'Total Amount'), (4, un_amt)]: ws2.cell(r, c, val)
            for c in range(1, 5): ws2.cell(r, c).fill = TOTAL_FILL; ws2.cell(r, c).font = BLACK_BOLD
            r += 1

            for c, h in enumerate(un_fields, 1):
                cell = ws2.cell(r, c, h); cell.fill, cell.font = EXC_HEADER_FILL, WHITE_BOLD
            r += 1

            for u in unmatched:
                ref = norm(u[s_i[schema['ref']]])
                amt = u[s_i[schema['amount']]]
                dt_val = u[s_i[schema['date']]] if schema.get('date') and schema['date'] in s_i else ''
                cur_val = u[s_i[schema['currency']]] if schema.get('currency') and schema['currency'] in s_i else cur
                ws2.cell(r, 1, ref); ws2.cell(r, 2, dt_val); ws2.cell(r, 3, amt)
                ws2.cell(r, 4, cur_val); ws2.cell(r, 5, str(sec['merchant_code']))
                ws2.cell(r, 6, sec['gw']); ws2.cell(r, 7, merchant)
                for c in range(1, 8): ws2.cell(r, c).fill = MISSING_FILL
                r += 1

            for c, val in [(1, 'Missing Count'), (2, len(unmatched)),
                           (3, 'Total Amount'), (4, un_amt)]: ws2.cell(r, c, val)
            for c in range(1, 5): ws2.cell(r, c).fill = TOTAL_FILL; ws2.cell(r, c).font = BLACK_BOLD
            r += 2
    autosize(ws2, 7)

def build_settlement_sheets(wb, sections, date_range):
    label = date_label(date_range)
    for sec in sorted(sections, key=lambda x: (x['currency'], x['merchant'])):
        cur, merchant, gw = sec['currency'], sec['merchant'], sec['gw']
        s_headers, schema = sec['headers'], sec['schema']
        end_date = date_range[1] if isinstance(date_range, tuple) else date_range
        short = merchant.replace(' Limited', '').replace(' Ltd', '') \
                         .replace(' Outfitters', '').replace(' Technology', '') \
                         .replace(' Fashions', '')[:18]
        sheet_name = f'{short} {cur} {end_date.strftime("%d%b").upper()}'[:31]
        ws = wb.create_sheet(sheet_name)
        r, ncols = 1, len(s_headers) + 6
        title = f'{merchant} {cur} \u2013 Settlement {label} ({gw})'
        write_title_row(ws, r, ncols, title); r += 1

        settle_total = sum(m[0][m[3][m[4]['amount']]] for m in sec['matched'])
        pelpay_total = sum(m[1]['amount'] for m in sec['matched'])
        diff_total = sum(m[2] for m in sec['matched'])
        write_total_row(ws, r, 6)
        for c, val in [(1, 'Settlement Total'), (2, settle_total),
                       (3, 'Pelpay Total'), (4, pelpay_total),
                       (5, 'Difference Total'), (6, diff_total)]: ws.cell(r, c, val)
        r += 1

        append_h = ['pelpay_transaction_date', 'pelpay_transaction_amount',
                     'pelpay_status', 'pelpay_merchant_reference',
                     'pelpay_payment_reference', 'settlement_minus_pelpay_difference']
        write_header_row(ws, r, s_headers + append_h + ['note']); r += 1

        matched_map = {norm(m[0][m[3][m[4]['ref']]]): m for m in sec['matched']}
        for sr in sec['rows']:
            ref = norm(sr[s_headers.index(schema['ref'])])
            for c, v in enumerate(sr, 1): ws.cell(r, c, v)
            m = matched_map.get(ref)
            if m:
                _, p, diff, _, _ = m
                for j, v in enumerate([p['dt'], p['amount'], p['status'],
                                        p['merchant_ref'], p['payment_ref'], diff]):
                    cell = ws.cell(r, len(s_headers) + 1 + j, v); cell.fill = MATCH_FILL
            else:
                for j in range(len(append_h)):
                    ws.cell(r, len(s_headers) + 1 + j).fill = MISSING_FILL
            r += 1

        write_total_row(ws, r, 6)
        for c, val in [(1, 'Settlement Total'), (2, settle_total),
                       (3, 'Pelpay Total'), (4, pelpay_total),
                       (5, 'Difference Total'), (6, diff_total)]: ws.cell(r, c, val)
        autosize(ws, ncols)


# ── Public API ───────────────────────────────────────────
def run(pelpay_path, settlement_files, settlement_date, output_path, schemas=None):
    """
    pelpay_path       — path to Pelpay combined .xlsx
    settlement_files  — list of (gateway_name: str, currency: str, file_path: str)
    settlement_date   — datetime.date or (start_date, end_date) tuple
    output_path       — where to save the result .xlsx
    schemas           — optional dict of gateway schemas (defaults to DEFAULT_SCHEMAS)
    Returns dict with keys: output_path, sheets, settle_rows, matched, exceptions
    """
    date_range = settlement_date if isinstance(settlement_date, tuple) else (settlement_date, settlement_date)
    pel_by_ref = load_pelpay(pelpay_path)
    sections = load_settlements(pel_by_ref, settlement_files, date_range, schemas)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    all_merchant_cur = sorted(set((s['merchant'], s['currency']) for s in sections))

    build_summary(wb_out, sections, date_range, pel_by_ref)
    build_received(wb_out, sections, date_range, pel_by_ref, all_merchant_cur)
    build_missing_sheets(wb_out, sections, date_range, all_merchant_cur)
    build_settlement_sheets(wb_out, sections, date_range)

    wb_out.save(output_path)
    sheet_names = list(wb_out.sheetnames)
    wb_out.close()

    wb2 = openpyxl.load_workbook(output_path)
    apply_formatting(wb2)
    wb2.save(output_path)

    total_settle = sum(len(s['rows']) for s in sections)
    total_matched = sum(len(s['matched']) for s in sections)
    total_exc = sum(len(s['exceptions']) for s in sections)

    return {
        'output_path': output_path,
        'sheets': sheet_names,
        'settle_rows': total_settle,
        'matched': total_matched,
        'exceptions': total_exc,
    }
